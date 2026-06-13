"""Excel (xlsx) eksport generatori — openpyxl asosida.

Ro'yxatlarni (buyurtmalar, qarzdorlik, moliya) chiroyli, formatlangan
Excel fayllarga chiqaradi. Sarlavha brend rangida, summalar raqam formatida,
ustun kengliklari avtomatik, sarlavha qatori muzlatilgan + avtofiltr.
"""
from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Callable, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Brend ranglari (frontend bilan mos)
_HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_ZEBRA_FILL = PatternFill("solid", fgColor="F8FAFC")
_THIN = Side(style="thin", color="E2E8F0")
_BORDER = Border(bottom=_THIN)

_MONEY_FMT = "#,##0 \"so'm\""
_DATE_FMT = "DD.MM.YYYY"


class Col:
    """Ustun ta'rifi: sarlavha, qiymat oluvchi va format turi."""
    def __init__(self, header: str, getter: Callable[[Any], Any],
                 kind: str = "text", width: Optional[int] = None):
        self.header = header
        self.getter = getter
        self.kind = kind  # text | money | date | int
        self.width = width


def _coerce(value: Any, kind: str) -> Any:
    if value is None:
        return None
    if kind == "money":
        try:
            return float(Decimal(str(value)))
        except Exception:
            return None
    if kind == "int":
        try:
            return int(value)
        except Exception:
            return value
    if kind == "date":
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day)
        return value
    return str(value)


def build_xlsx(sheet_title: str, columns: Sequence[Col], rows: Sequence[Any]) -> bytes:
    """Ustun ta'riflari va qatorlardan formatlangan xlsx bytes qaytaradi."""
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Hisobot")[:31]

    # Sarlavha qatori
    for ci, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=ci, value=col.header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ma'lumot qatorlari
    for ri, obj in enumerate(rows, start=2):
        for ci, col in enumerate(columns, start=1):
            raw = col.getter(obj)
            value = _coerce(raw, col.kind)
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.border = _BORDER
            if col.kind == "money":
                cell.number_format = _MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
            elif col.kind == "int":
                cell.alignment = Alignment(horizontal="center")
            elif col.kind == "date":
                cell.number_format = _DATE_FMT
                cell.alignment = Alignment(horizontal="center")
            if ri % 2 == 0:
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = _ZEBRA_FILL

    # Ustun kengliklari (avtomatik, sarlavha va qiymatlar bo'yicha)
    for ci, col in enumerate(columns, start=1):
        if col.width:
            width = col.width
        else:
            max_len = len(str(col.header))
            for obj in rows[:200]:  # katta ro'yxatda faqat boshini o'lchaymiz
                v = col.getter(obj)
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            width = min(max(max_len + 2, 10), 40)
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Muzlatish + avtofiltr
    ws.freeze_panes = "A2"
    if columns:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, len(rows) + 1)}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
#  Statuslar — o'zbekcha yorliqlar
# --------------------------------------------------------------------------- #
_ORDER_STATUS = {
    "new": "Navbatda", "ready": "Tayyor bo'ldi",
    "delivered": "Yetkazildi", "rejected": "Rad etildi",
}


def order_status_label(status: str) -> str:
    return _ORDER_STATUS.get(status, status or "—")


# --------------------------------------------------------------------------- #
#  Buyurtmalar ro'yxati
# --------------------------------------------------------------------------- #
def orders_workbook(orders: Sequence[Any]) -> bytes:
    columns = [
        Col("Kod", lambda o: o.code, width=16),
        Col("Sana", lambda o: o.order_date, kind="date"),
        Col("Mijoz", lambda o: (o.customer.full_name if o.customer else "—"), width=28),
        Col("Telefon", lambda o: (o.customer.phone if o.customer else None), width=18),
        Col("Viloyat", lambda o: (o.customer.region if o.customer else None), width=18),
        Col("Holat", lambda o: order_status_label(o.status), width=14),
        Col("Mahsulot", lambda o: len(o.items), kind="int"),
        Col("Jami", lambda o: o.items_total_uzs, kind="money"),
        Col("To'langan", lambda o: o.paid_uzs, kind="money"),
        Col("Qoldiq", lambda o: o.balance_uzs, kind="money"),
    ]
    return build_xlsx("Buyurtmalar", columns, list(orders))


# --------------------------------------------------------------------------- #
#  Qarzdorlik (receivables)
# --------------------------------------------------------------------------- #
def receivables_workbook(items: Sequence[dict]) -> bytes:
    columns = [
        Col("Mijoz", lambda r: r.get("customer"), width=28),
        Col("Telefon", lambda r: r.get("phone"), width=18),
        Col("Buyurtma", lambda r: r.get("code"), width=16),
        Col("Sana", lambda r: r.get("order_date"), kind="date"),
        Col("Kun", lambda r: r.get("days"), kind="int"),
        Col("Jami", lambda r: r.get("total_uzs"), kind="money"),
        Col("To'langan", lambda r: r.get("paid_uzs"), kind="money"),
        Col("Qoldiq", lambda r: r.get("balance_uzs"), kind="money"),
    ]
    return build_xlsx("Qarzdorlik", columns, list(items))
